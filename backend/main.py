from __future__ import annotations

import io
import logging
import os
from typing import Optional

logger = logging.getLogger(__name__)

import pandas as pd
from fastapi import FastAPI, File, HTTPException, Query, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
from pydantic import BaseModel

from database import (
    get_connection, init_db, seed_data,
    db_exec, db_exec_many, db_scalar, insert_returning_id, month_expr,
    USE_POSTGRES,
)

app = FastAPI(title="Fazenda Nutrition Dashboard API", version="3.0.0")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)


_db_ready = False


@app.on_event("startup")
async def on_startup():
    global _db_ready
    try:
        init_db()
        seed_data()
        _db_ready = True
    except Exception as exc:
        logger.error("Falha ao inicializar o banco de dados: %s", exc)
        logger.error("Verifique a variável DATABASE_URL no painel do Render.")


# ──────────────────────────────────────────────
# Health
# ──────────────────────────────────────────────

@app.get("/api/health")
def health():
    if not _db_ready:
        return {"status": "error", "db": "unavailable",
                "detail": "Banco não inicializado — verifique DATABASE_URL no Render"}
    try:
        conn = get_connection()
        db_exec(conn, "SELECT 1")
        conn.close()
        return {"status": "ok", "db": "postgresql" if USE_POSTGRES else "sqlite"}
    except Exception as exc:
        logger.error("Health check DB query failed: %s", exc)
        return {"status": "error", "db": "unavailable", "detail": str(exc)}


# ──────────────────────────────────────────────
# WHERE builder
# ──────────────────────────────────────────────

def _build_where(
    data_inicio: Optional[str],
    data_fim:    Optional[str],
    lotes_str:   Optional[str],
    fazenda:     Optional[str] = None,
) -> tuple[str, list]:
    clauses, params = [], []
    if fazenda:
        clauses.append("fazenda = ?")
        params.append(fazenda)
    if data_inicio:
        clauses.append("data_registro >= ?")
        params.append(data_inicio)
    if data_fim:
        clauses.append("data_registro <= ?")
        params.append(data_fim)
    if lotes_str:
        lote_list = [l.strip() for l in lotes_str.split(",") if l.strip()]
        if lote_list:
            ph = ",".join("?" * len(lote_list))
            clauses.append(f"lote IN ({ph})")
            params.extend(lote_list)
    where = ("WHERE " + " AND ".join(clauses)) if clauses else ""
    return where, params


# ──────────────────────────────────────────────
# Meta endpoints
# ──────────────────────────────────────────────

@app.get("/api/fazendas")
def get_fazendas():
    conn = get_connection()
    rows = db_exec(conn,
        "SELECT DISTINCT fazenda FROM producao_rebanho WHERE fazenda != '' ORDER BY fazenda"
    ).fetchall()
    conn.close()
    return {"fazendas": [r["fazenda"] for r in rows]}


@app.get("/api/lotes")
def get_lotes(fazenda: Optional[str] = Query(None)):
    conn = get_connection()
    if fazenda:
        rows = db_exec(conn,
            "SELECT DISTINCT lote FROM producao_rebanho WHERE fazenda = ? ORDER BY lote",
            [fazenda],
        ).fetchall()
    else:
        rows = db_exec(conn,
            "SELECT DISTINCT lote FROM producao_rebanho ORDER BY lote"
        ).fetchall()
    conn.close()
    return {"lotes": [r["lote"] for r in rows]}


@app.get("/api/dates")
def get_dates(fazenda: Optional[str] = Query(None)):
    conn = get_connection()
    if fazenda:
        row = db_exec(conn,
            "SELECT MIN(data_registro) AS min_d, MAX(data_registro) AS max_d "
            "FROM producao_rebanho WHERE fazenda = ?",
            [fazenda],
        ).fetchone()
    else:
        row = db_exec(conn,
            "SELECT MIN(data_registro) AS min_d, MAX(data_registro) AS max_d "
            "FROM producao_rebanho"
        ).fetchone()
    conn.close()
    return {
        "min_date": str(row["min_d"]) if row["min_d"] else None,
        "max_date": str(row["max_d"]) if row["max_d"] else None,
    }


# ──────────────────────────────────────────────
# Configurações — Meta de sobra
# ──────────────────────────────────────────────

class MetaSobraIn(BaseModel):
    fazenda: str
    lote: str
    meta_pct: float


@app.get("/api/config/meta-sobra")
def get_meta_sobra(fazenda: Optional[str] = Query(None)):
    conn = get_connection()
    if fazenda:
        rows = db_exec(conn,
            "SELECT fazenda, lote, meta_pct FROM meta_sobra WHERE fazenda = ? ORDER BY lote",
            [fazenda],
        ).fetchall()
    else:
        rows = db_exec(conn,
            "SELECT fazenda, lote, meta_pct FROM meta_sobra ORDER BY fazenda, lote"
        ).fetchall()
    conn.close()
    return {"data": [dict(r) for r in rows]}


@app.post("/api/config/meta-sobra")
def upsert_meta_sobra(body: MetaSobraIn):
    conn = get_connection()
    if USE_POSTGRES:
        db_exec(conn,
            """INSERT INTO meta_sobra (fazenda, lote, meta_pct) VALUES (?, ?, ?)
               ON CONFLICT (fazenda, lote) DO UPDATE SET meta_pct = EXCLUDED.meta_pct""",
            [body.fazenda, body.lote, body.meta_pct],
        )
    else:
        db_exec(conn,
            "INSERT OR REPLACE INTO meta_sobra (fazenda, lote, meta_pct) VALUES (?, ?, ?)",
            [body.fazenda, body.lote, body.meta_pct],
        )
    conn.commit()
    conn.close()
    return {"ok": True}


@app.delete("/api/config/meta-sobra")
def delete_meta_sobra(fazenda: str = Query(...), lote: str = Query(...)):
    conn = get_connection()
    db_exec(conn, "DELETE FROM meta_sobra WHERE fazenda = ? AND lote = ?", [fazenda, lote])
    conn.commit()
    conn.close()
    return {"ok": True}


# ──────────────────────────────────────────────
# Dashboard data
# ──────────────────────────────────────────────

@app.get("/api/dashboard/data")
def dashboard_data(
    data_inicio: Optional[str] = Query(None),
    data_fim:    Optional[str] = Query(None),
    lotes:       Optional[str] = Query(None),
    fazenda:     Optional[str] = Query(None),
):
    where, params = _build_where(data_inicio, data_fim, lotes, fazenda)
    conn = get_connection()
    rows = db_exec(conn,
        f"""SELECT data_registro, fazenda, lote, num_vacas,
                   producao_leite_total, leite_por_vaca,
                   consumo_ms_total, consumo_ms_vaca,
                   percentual_forragem, eficiencia_alimentar,
                   pct_ms_forragem, pct_ms_dieta,
                   qtd_dieta_fornecida, qtd_sobra_dieta, pct_sobra
            FROM producao_rebanho
            {where}
            ORDER BY data_registro, lote""",
        params,
    ).fetchall()
    conn.close()
    return {"data": [dict(r) for r in rows]}


@app.get("/api/dashboard/kpis")
def dashboard_kpis(
    data_inicio: Optional[str] = Query(None),
    data_fim:    Optional[str] = Query(None),
    lotes:       Optional[str] = Query(None),
    fazenda:     Optional[str] = Query(None),
):
    where, params = _build_where(data_inicio, data_fim, lotes, fazenda)
    conn = get_connection()

    last_row = db_exec(conn,
        f"SELECT MAX(data_registro) AS max_d FROM producao_rebanho {where}", params
    ).fetchone()
    last_date = db_scalar(last_row, 'max_d')
    if last_date:
        last_date = str(last_date)

    if not last_date:
        conn.close()
        return {
            "total_producao": 0, "total_vacas": 0,
            "eficiencia_ponderada": None, "avg_forragem": None,
            "leite_vaca_rebanho": None, "ms_vaca_rebanho": None,
            "avg_pct_ms_forragem": None, "avg_pct_ms_dieta": None,
            "avg_pct_sobra": None,
        }

    where2, params2 = _build_where(last_date, last_date, lotes, fazenda)
    totals = db_exec(conn,
        f"""SELECT
               SUM(producao_leite_total)                                              AS total_prod,
               SUM(num_vacas)                                                         AS total_vacas,
               SUM(consumo_ms_total)                                                  AS total_ms,
               AVG(percentual_forragem)                                               AS avg_forr,
               SUM(pct_ms_forragem * num_vacas) /
                   NULLIF(SUM(CASE WHEN pct_ms_forragem IS NOT NULL THEN num_vacas ELSE 0 END), 0)
                                                                                      AS pct_ms_forr_pond,
               SUM(pct_ms_dieta * num_vacas) /
                   NULLIF(SUM(CASE WHEN pct_ms_dieta IS NOT NULL THEN num_vacas ELSE 0 END), 0)
                                                                                      AS pct_ms_dieta_pond,
               AVG(pct_sobra)                                                         AS avg_pct_sobra
            FROM producao_rebanho {where2}""",
        params2,
    ).fetchone()
    conn.close()

    def _f(v):
        return float(v) if v is not None else None

    total_prod  = _f(totals["total_prod"])  or 0
    total_ms    = _f(totals["total_ms"])
    total_vacas = int(totals["total_vacas"] or 0)
    avg_forr    = _f(totals["avg_forr"])
    pf          = _f(totals["pct_ms_forr_pond"])
    pd_         = _f(totals["pct_ms_dieta_pond"])
    avg_sobra   = _f(totals["avg_pct_sobra"])

    return {
        "total_producao":       round(total_prod, 1),
        "total_vacas":          total_vacas,
        "eficiencia_ponderada": round(total_prod / total_ms, 4) if total_ms else None,
        "avg_forragem":         round(avg_forr, 1) if avg_forr is not None else None,
        "leite_vaca_rebanho":   round(total_prod / total_vacas, 2) if total_vacas else None,
        "ms_vaca_rebanho":      round(total_ms / total_vacas, 2) if total_ms and total_vacas else None,
        "avg_pct_ms_forragem":  round(pf, 1)  if pf  is not None else None,
        "avg_pct_ms_dieta":     round(pd_, 1) if pd_ is not None else None,
        "avg_pct_sobra":        round(avg_sobra, 1) if avg_sobra is not None else None,
        "ultima_data":          last_date,
    }


def _batch_sql_pg(where):
    return f"""SELECT
           lote,
           ROUND(AVG(num_vacas)::numeric, 1)                                                     AS avg_vacas,
           ROUND((SUM(producao_leite_total)/NULLIF(SUM(num_vacas),0))::numeric, 2)               AS leite_vaca_pond,
           ROUND((SUM(consumo_ms_total)/NULLIF(SUM(num_vacas),0))::numeric, 2)                   AS ms_vaca_pond,
           ROUND((SUM(producao_leite_total)/NULLIF(SUM(consumo_ms_total),0))::numeric, 4)        AS eficiencia_pond,
           ROUND(AVG(percentual_forragem)::numeric, 1)                                           AS avg_forragem,
           ROUND(AVG(pct_ms_forragem)::numeric, 1)                                               AS avg_pct_ms_forragem,
           ROUND(AVG(pct_ms_dieta)::numeric, 1)                                                  AS avg_pct_ms_dieta,
           ROUND(AVG(pct_sobra)::numeric, 2)                                                     AS avg_pct_sobra,
           COUNT(*)                                                                               AS num_registros
        FROM producao_rebanho
        {where}
        GROUP BY lote
        ORDER BY eficiencia_pond DESC NULLS LAST"""


def _batch_sql_sq(where):
    return f"""SELECT
           lote,
           ROUND(AVG(num_vacas), 1)                                                              AS avg_vacas,
           ROUND(SUM(producao_leite_total)/NULLIF(SUM(num_vacas),0), 2)                          AS leite_vaca_pond,
           ROUND(SUM(consumo_ms_total)/NULLIF(SUM(num_vacas),0), 2)                              AS ms_vaca_pond,
           ROUND(SUM(producao_leite_total)/NULLIF(SUM(consumo_ms_total),0), 4)                   AS eficiencia_pond,
           ROUND(AVG(percentual_forragem), 1)                                                    AS avg_forragem,
           ROUND(AVG(pct_ms_forragem), 1)                                                        AS avg_pct_ms_forragem,
           ROUND(AVG(pct_ms_dieta), 1)                                                           AS avg_pct_ms_dieta,
           ROUND(AVG(pct_sobra), 2)                                                              AS avg_pct_sobra,
           COUNT(*)                                                                               AS num_registros
        FROM producao_rebanho
        {where}
        GROUP BY lote
        ORDER BY eficiencia_pond DESC"""


@app.get("/api/dashboard/batch-summary")
def batch_summary(
    data_inicio: Optional[str] = Query(None),
    data_fim:    Optional[str] = Query(None),
    lotes:       Optional[str] = Query(None),
    fazenda:     Optional[str] = Query(None),
):
    where, params = _build_where(data_inicio, data_fim, lotes, fazenda)
    conn = get_connection()
    sql = _batch_sql_pg(where) if USE_POSTGRES else _batch_sql_sq(where)
    rows = db_exec(conn, sql, params).fetchall()
    conn.close()
    return {"data": [dict(r) for r in rows]}


@app.get("/api/dashboard/monthly")
def dashboard_monthly(
    data_inicio: Optional[str] = Query(None),
    data_fim:    Optional[str] = Query(None),
    lotes:       Optional[str] = Query(None),
    fazenda:     Optional[str] = Query(None),
):
    where, params = _build_where(data_inicio, data_fim, lotes, fazenda)
    mes = month_expr("data_registro")
    conn = get_connection()
    if USE_POSTGRES:
        sql = f"""SELECT
               {mes}                                                                              AS mes,
               lote,
               ROUND(AVG(num_vacas)::numeric, 1)                                                 AS avg_vacas,
               ROUND((SUM(producao_leite_total)/NULLIF(SUM(num_vacas),0))::numeric, 2)           AS leite_vaca_pond,
               ROUND((SUM(consumo_ms_total)/NULLIF(SUM(num_vacas),0))::numeric, 2)               AS ms_vaca_pond,
               ROUND((SUM(producao_leite_total)/NULLIF(SUM(consumo_ms_total),0))::numeric, 4)    AS eficiencia_pond,
               ROUND(AVG(percentual_forragem)::numeric, 1)                                       AS avg_forragem,
               ROUND(AVG(pct_ms_forragem)::numeric, 1)                                           AS avg_pct_ms_forragem,
               ROUND(AVG(pct_ms_dieta)::numeric, 1)                                              AS avg_pct_ms_dieta,
               ROUND(AVG(pct_sobra)::numeric, 2)                                                 AS avg_pct_sobra,
               COUNT(*)                                                                           AS num_registros
            FROM producao_rebanho
            {where}
            GROUP BY {mes}, lote
            ORDER BY {mes}, lote"""
    else:
        sql = f"""SELECT
               {mes}                                                                              AS mes,
               lote,
               ROUND(AVG(num_vacas), 1)                                                          AS avg_vacas,
               ROUND(SUM(producao_leite_total)/NULLIF(SUM(num_vacas),0), 2)                      AS leite_vaca_pond,
               ROUND(SUM(consumo_ms_total)/NULLIF(SUM(num_vacas),0), 2)                          AS ms_vaca_pond,
               ROUND(SUM(producao_leite_total)/NULLIF(SUM(consumo_ms_total),0), 4)               AS eficiencia_pond,
               ROUND(AVG(percentual_forragem), 1)                                                AS avg_forragem,
               ROUND(AVG(pct_ms_forragem), 1)                                                    AS avg_pct_ms_forragem,
               ROUND(AVG(pct_ms_dieta), 1)                                                       AS avg_pct_ms_dieta,
               ROUND(AVG(pct_sobra), 2)                                                          AS avg_pct_sobra,
               COUNT(*)                                                                           AS num_registros
            FROM producao_rebanho
            {where}
            GROUP BY {mes}, lote
            ORDER BY {mes}, lote"""
    rows = db_exec(conn, sql, params).fetchall()
    conn.close()
    return {"data": [dict(r) for r in rows]}


@app.get("/api/dashboard/raw")
def dashboard_raw(
    data_inicio: Optional[str] = Query(None),
    data_fim:    Optional[str] = Query(None),
    lotes:       Optional[str] = Query(None),
    fazenda:     Optional[str] = Query(None),
):
    where, params = _build_where(data_inicio, data_fim, lotes, fazenda)
    conn = get_connection()
    rows = db_exec(conn,
        f"""SELECT
               pr.id, pr.fazenda, pr.data_registro, pr.lote, pr.num_vacas,
               pr.producao_leite_total, pr.leite_por_vaca,
               pr.consumo_ms_total, pr.consumo_ms_vaca,
               pr.percentual_forragem, pr.eficiencia_alimentar,
               pr.pct_ms_forragem, pr.pct_ms_dieta,
               pr.qtd_dieta_fornecida, pr.qtd_sobra_dieta, pr.pct_sobra,
               pr.created_at,
               u.filename AS upload_filename
            FROM producao_rebanho pr
            LEFT JOIN uploads u ON pr.upload_id = u.id
            {where}
            ORDER BY pr.data_registro DESC, pr.lote""",
        params,
    ).fetchall()
    conn.close()
    return {"data": [dict(r) for r in rows]}


# ──────────────────────────────────────────────
# Upload
# ──────────────────────────────────────────────

# Apenas os 4 identificadores são obrigatórios — o resto é opcional
REQUIRED_COLS = {"fazenda", "data", "lote", "num_vacas"}

# Colunas opcionais reconhecidas
OPTIONAL_COLS = {
    "producao_leite_total", "leite_por_vaca",
    "qtd_dieta_fornecida", "pct_ms_dieta", "qtd_sobra_dieta",
    "percentual_forragem", "pct_ms_forragem",
}


def _nullable(val) -> Optional[float]:
    if val is None:
        return None
    try:
        import math
        if math.isnan(float(val)):
            return None
        return float(val)
    except (TypeError, ValueError):
        return None


@app.post("/api/upload")
async def upload_excel(file: UploadFile = File(...)):
    if not file.filename.endswith((".xlsx", ".xls", ".csv")):
        raise HTTPException(400, "Apenas .xlsx, .xls ou .csv são aceitos.")

    content = await file.read()
    try:
        if file.filename.endswith(".csv"):
            enc = "utf-8-sig"
            try:
                content.decode("utf-8-sig")
            except UnicodeDecodeError:
                enc = "latin-1"
            df = pd.read_csv(io.BytesIO(content), encoding=enc, sep=None, engine="python")
        else:
            df = pd.read_excel(io.BytesIO(content))
    except Exception as e:
        raise HTTPException(400, f"Erro ao ler arquivo: {e}")

    df.columns = [c.strip().lower().replace(" ", "_") for c in df.columns]

    # Aceita num_cabecas como alias de num_vacas
    if "num_cabecas" in df.columns and "num_vacas" not in df.columns:
        df.rename(columns={"num_cabecas": "num_vacas"}, inplace=True)

    missing = REQUIRED_COLS - set(df.columns)
    if missing:
        raise HTTPException(400,
            f"Colunas obrigatórias ausentes: {', '.join(sorted(missing))}. "
            "Baixe o template para ver o formato correto.")

    # Remover linhas sem data válida (ex: linha de descrição do template)
    df = df[pd.to_datetime(df["data"], errors="coerce").notna()].copy()
    if df.empty:
        raise HTTPException(400,
            "Nenhuma linha com data válida encontrada. "
            "Verifique o formato da coluna 'data' (DD/MM/AAAA ou AAAA-MM-DD).")

    df["data"] = pd.to_datetime(df["data"], dayfirst=True).dt.date
    df["num_vacas"] = pd.to_numeric(df["num_vacas"], errors="coerce").fillna(0).astype(int)

    # Garantir que colunas opcionais existam como NaN
    for col in OPTIONAL_COLS:
        if col not in df.columns:
            df[col] = float("nan")
        else:
            df[col] = pd.to_numeric(df[col], errors="coerce")

    # ── Derivar leite ─────────────────────────────────────────────────────
    has_total = df["producao_leite_total"].notna()
    has_avg   = df["leite_por_vaca"].notna()
    nv        = df["num_vacas"].replace(0, float("nan"))

    mask_t = has_total & ~has_avg
    mask_a = ~has_total & has_avg
    df.loc[mask_t, "leite_por_vaca"]        = (df.loc[mask_t, "producao_leite_total"] / nv.loc[mask_t]).round(4)
    df.loc[mask_a, "producao_leite_total"]  = (df.loc[mask_a, "leite_por_vaca"] * df.loc[mask_a, "num_vacas"]).round(1)

    # ── Derivar CMS ───────────────────────────────────────────────────────
    df["consumo_ms_total"] = float("nan")
    df["consumo_ms_vaca"]  = float("nan")

    mask_cms = (
        df["qtd_dieta_fornecida"].notna() &
        df["qtd_sobra_dieta"].notna() &
        df["pct_ms_dieta"].notna() &
        (df["num_vacas"] > 0)
    )
    if mask_cms.any():
        consumido = df.loc[mask_cms, "qtd_dieta_fornecida"] - df.loc[mask_cms, "qtd_sobra_dieta"]
        df.loc[mask_cms, "consumo_ms_total"] = (consumido * df.loc[mask_cms, "pct_ms_dieta"] / 100).round(2)
        df.loc[mask_cms, "consumo_ms_vaca"]  = (df.loc[mask_cms, "consumo_ms_total"] / df.loc[mask_cms, "num_vacas"]).round(4)

    # ── Derivar % sobra ───────────────────────────────────────────────────
    df["pct_sobra"] = float("nan")
    mask_sobra = (
        df["qtd_dieta_fornecida"].notna() &
        df["qtd_sobra_dieta"].notna() &
        (df["qtd_dieta_fornecida"] > 0)
    )
    if mask_sobra.any():
        df.loc[mask_sobra, "pct_sobra"] = (
            df.loc[mask_sobra, "qtd_sobra_dieta"] / df.loc[mask_sobra, "qtd_dieta_fornecida"] * 100
        ).round(2)

    # ── Derivar eficiência ────────────────────────────────────────────────
    df["eficiencia_alimentar"] = float("nan")
    mask_ef = (
        df["leite_por_vaca"].notna() &
        df["consumo_ms_vaca"].notna() &
        (df["consumo_ms_vaca"] > 0)
    )
    if mask_ef.any():
        df.loc[mask_ef, "eficiencia_alimentar"] = (
            df.loc[mask_ef, "leite_por_vaca"] / df.loc[mask_ef, "consumo_ms_vaca"]
        ).round(4)

    conn = get_connection()
    upload_id = insert_returning_id(conn,
        "INSERT INTO uploads (filename, num_records) VALUES (?, ?)",
        (file.filename, len(df)),
    )

    records = [
        (
            upload_id,
            str(row["fazenda"]).strip(),
            row["data"].isoformat(),
            str(row["lote"]).strip(),
            int(row["num_vacas"]),
            _nullable(row.get("producao_leite_total")),
            _nullable(row.get("leite_por_vaca")),
            _nullable(row.get("consumo_ms_total")),
            _nullable(row.get("consumo_ms_vaca")),
            _nullable(row.get("percentual_forragem")),
            _nullable(row.get("eficiencia_alimentar")),
            _nullable(row.get("pct_ms_forragem")),
            _nullable(row.get("pct_ms_dieta")),
            _nullable(row.get("qtd_dieta_fornecida")),
            _nullable(row.get("qtd_sobra_dieta")),
            _nullable(row.get("pct_sobra")),
        )
        for _, row in df.iterrows()
    ]

    db_exec_many(conn,
        """INSERT INTO producao_rebanho
           (upload_id, fazenda, data_registro, lote, num_vacas,
            producao_leite_total, leite_por_vaca,
            consumo_ms_total, consumo_ms_vaca,
            percentual_forragem, eficiencia_alimentar,
            pct_ms_forragem, pct_ms_dieta,
            qtd_dieta_fornecida, qtd_sobra_dieta, pct_sobra)
           VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
        records,
    )
    conn.commit()
    conn.close()

    return {
        "message": "Upload realizado com sucesso!",
        "records_inserted": len(records),
        "fazendas": sorted({r[1] for r in records}),
    }


@app.get("/api/uploads")
def list_uploads():
    conn = get_connection()
    rows = db_exec(conn,
        "SELECT id, filename, uploaded_at, num_records, notes FROM uploads ORDER BY uploaded_at DESC"
    ).fetchall()
    conn.close()
    return {"data": [dict(r) for r in rows]}


# ──────────────────────────────────────────────
# Template download
# ──────────────────────────────────────────────

@app.get("/api/template")
def download_template(fmt: str = Query("xlsx", pattern="^(xlsx|csv)$")):
    col_keys = [
        "fazenda", "data", "lote", "num_vacas",
        "producao_leite_total", "leite_por_vaca",
        "qtd_dieta_fornecida", "pct_ms_dieta",
        "qtd_sobra_dieta", "percentual_forragem", "pct_ms_forragem",
    ]
    col_labels = [
        "Fazenda *", "Data *", "Lote *", "N Vacas *",
        "Prod. Leite Total kg", "Leite/Vaca kg",
        "Dieta Fornecida kg MN *", "% MS Dieta *",
        "Sobra kg MN", "% Forragem", "% MS Forragem",
    ]
    col_hints = [
        "Nome da fazenda",
        "DD/MM/AAAA ou AAAA-MM-DD",
        "Ex: TOP VACA, CB1...",
        "Número inteiro",
        "kg leite total do lote (ou use Leite/Vaca)",
        "kg leite por vaca/dia (ou use Prod. Total)",
        "kg de dieta fornecida (MN)",
        "% de matéria seca da dieta (ex: 47.0)",
        "kg de sobra (MN) — opcional",
        "% forragem na dieta (ex: 52.5) — opcional",
        "% MS da forragem — opcional",
    ]
    required = [True, True, True, True, False, False, True, True, False, False, False]

    examples = [
        ["Fazenda Sao Joao", "01/05/2026", "TOP VACA",  142, 5453, 38.4, 8920, 49.6,  580, 52.1, 38.2],
        ["Fazenda Sao Joao", "01/05/2026", "TOP NOV",    98, 3205, 32.7, 5910, 47.8,  370, 54.8, 36.9],
        ["Fazenda Sao Joao", "01/05/2026", "CB1",       165, 4653, 28.2, 9150, 46.2,  620, 56.3, 35.4],
        ["Fazenda Sao Joao", "01/05/2026", "CB2",       134, 3296, 24.6, 7280, 44.7,  510, 58.9, 34.1],
        ["Fazenda Sao Joao", "01/05/2026", "CB4",        89, 1762, 19.8, 4650, 43.1,  340, 61.4, 32.8],
        ["Fazenda Sao Joao", "01/05/2026", "POS PARTO",  42, 1474, 35.1, 2540, 50.4,  160, 50.8, 39.1],
        ["Fazenda Sao Joao", "02/05/2026", "TOP VACA",  142, 5481, 38.6, 8950, 49.5,  575, 52.0, 38.1],
        ["Fazenda Sao Joao", "02/05/2026", "TOP NOV",    98, 3224, 32.9, 5940, 47.7,  368, 54.7, 36.8],
        ["Fazenda Sao Joao", "02/05/2026", "CB1",       165, 4686, 28.4, 9200, 46.1,  628, 56.2, 35.3],
        ["Fazenda Sao Joao", "02/05/2026", "CB2",       134, 3243, 24.2, 7310, 44.6,  505, 59.0, 34.0],
        ["Fazenda Sao Joao", "02/05/2026", "CB4",        89, 1744, 19.6, 4680, 43.0,  342, 61.5, 32.7],
        ["Fazenda Sao Joao", "02/05/2026", "POS PARTO",  42, 1450, 34.5, 2520, 50.2,  158, 51.0, 39.0],
    ]

    if fmt == "csv":
        import csv
        from io import StringIO
        buf = StringIO()
        w = csv.writer(buf)
        w.writerow(col_keys)
        for ex in examples:
            w.writerow(ex)
        buf.seek(0)
        return StreamingResponse(
            iter([buf.getvalue().encode("utf-8-sig")]),
            media_type="text/csv",
            headers={"Content-Disposition": "attachment; filename=template_fazenda.csv"},
        )

    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()

    # ── Sheet 1: Dados ──────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Dados"
    n_cols = len(col_keys)
    last_col = get_column_letter(n_cols)

    def fill(hex_color):
        return PatternFill("solid", fgColor=hex_color)

    def border():
        t = Side(style="thin", color="CCCCCC")
        return Border(left=t, right=t, top=t, bottom=t)

    def font(bold=False, italic=False, size=10, color="1A1F1A"):
        return Font(bold=bold, italic=italic, size=size, color=color, name="Calibri")

    def align(h="left", v="center", wrap=False):
        return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

    # Row 1 — title banner
    ws.merge_cells(f"A1:{last_col}1")
    c = ws["A1"]
    c.value = "Planilha de Importacao — Dashboard Fazenda"
    c.font = font(bold=True, size=14, color="FFFFFF")
    c.fill = fill("1A6B3A")
    c.alignment = align("center")
    ws.row_dimensions[1].height = 30

    # Row 2 — instruction note
    ws.merge_cells(f"A2:{last_col}2")
    c = ws["A2"]
    c.value = (
        "Preencha a partir da linha 6. Colunas com * sao obrigatorias. "
        "Use producao_leite_total OU leite_por_vaca (nao precisa de ambos). "
        "As linhas 6 a 17 sao apenas exemplos — apague-as antes de enviar."
    )
    c.font = font(italic=True, size=10, color="2F6B3D")
    c.fill = fill("E8F2EA")
    c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)
    ws.row_dimensions[2].height = 28

    # Row 3 — friendly column labels
    for ci, (label, req) in enumerate(zip(col_labels, required), 1):
        c = ws.cell(row=3, column=ci, value=label)
        c.font = font(bold=True, size=10, color="FFFFFF")
        c.fill = fill("1A6B3A") if req else fill("6B7568")
        c.alignment = align("center", wrap=True)
        c.border = border()
    ws.row_dimensions[3].height = 32

    # Row 4 — technical key names
    for ci, key in enumerate(col_keys, 1):
        c = ws.cell(row=4, column=ci, value=key)
        c.font = font(italic=True, size=8, color="555555")
        c.fill = fill("E8F2EA")
        c.alignment = align("center", wrap=True)
        c.border = border()
    ws.row_dimensions[4].height = 16

    # Row 5 — per-column hints
    for ci, hint in enumerate(col_hints, 1):
        c = ws.cell(row=5, column=ci, value=hint)
        c.font = font(italic=True, size=9, color="3A4438")
        c.fill = fill("F0F7F1")
        c.alignment = align("center", wrap=True)
        c.border = border()
    ws.row_dimensions[5].height = 36

    # Rows 6–17 — example data
    for ri, row_data in enumerate(examples, 6):
        for ci, val in enumerate(row_data, 1):
            c = ws.cell(row=ri, column=ci, value=val)
            c.fill = fill("FFFDE7")
            c.border = border()
            c.font = font(bold=(ci == 3), size=10)
            c.alignment = align("left" if ci <= 3 else "right")
        ws.row_dimensions[ri].height = 18

    # Rows 18–47 — empty input rows
    for ri in range(18, 48):
        for ci in range(1, n_cols + 1):
            c = ws.cell(row=ri, column=ci)
            c.fill = fill("FAFFF9")
            c.border = border()
            c.alignment = align("left" if ci <= 3 else "right")
        ws.row_dimensions[ri].height = 18

    # Column widths
    for ci, w in enumerate([22, 16, 14, 10, 18, 14, 20, 14, 14, 14, 16], 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    ws.freeze_panes = "A6"

    # ── Sheet 2: Instrucoes ─────────────────────────────────────────────────
    wi = wb.create_sheet("Instrucoes")
    wi.column_dimensions["A"].width = 26
    wi.column_dimensions["B"].width = 72

    def wi_header(ri, text, bg):
        wi.merge_cells(f"A{ri}:B{ri}")
        c = wi.cell(row=ri, column=1, value=text)
        c.font = font(bold=True, size=11, color="FFFFFF")
        c.fill = fill(bg)
        c.alignment = align("left", wrap=True)
        wi.row_dimensions[ri].height = 22

    def wi_row(ri, label, desc, bg_a, bg_b=None):
        ca = wi.cell(row=ri, column=1, value=label)
        ca.font = font(bold=bool(desc is None), size=10)
        ca.fill = fill(bg_a)
        ca.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        if desc is not None:
            cb = wi.cell(row=ri, column=2, value=desc)
            cb.font = font(size=10)
            cb.fill = fill(bg_b or bg_a)
            cb.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        wi.row_dimensions[ri].height = 20

    r = 1
    wi_header(r, "GUIA DE PREENCHIMENTO — Dashboard Fazenda", "1A6B3A"); r += 1
    wi.row_dimensions[r].height = 6; r += 1

    wi_header(r, "CAMPOS OBRIGATORIOS  (marcados com *)", "1A6B3A"); r += 1
    for label, desc in [
        ("fazenda",              "Nome da fazenda. Use exatamente o mesmo nome em todos os registros."),
        ("data",                 "Data do registro. Formatos aceitos: DD/MM/AAAA (ex: 01/05/2026) ou AAAA-MM-DD."),
        ("lote",                 "Nome do lote. Exemplos: TOP VACA, TOP NOV, CB1, CB2, CB4, POS PARTO."),
        ("num_vacas",            "Total de vacas no lote nessa data. Deve ser um numero inteiro (ex: 142)."),
        ("qtd_dieta_fornecida",  "Quantidade total de dieta fornecida ao lote em kg de Materia Natural (MN)."),
        ("pct_ms_dieta",         "Percentual de materia seca da dieta total (ex: 47.0). Usado para calcular o CMS."),
    ]:
        wi_row(r, label, desc, "E8F2EA"); r += 1

    wi.row_dimensions[r].height = 6; r += 1
    wi_header(r, "CAMPOS OPCIONAIS  (use um ou mais conforme disponibilidade)", "6B7568"); r += 1
    for label, desc in [
        ("producao_leite_total", "Producao total de leite do lote no dia (kg). Use este OU leite_por_vaca — nao precisa de ambos."),
        ("leite_por_vaca",       "Producao media de leite por vaca no dia (kg). Use este OU producao_leite_total."),
        ("qtd_sobra_dieta",      "Quantidade de sobra da dieta em kg MN. Necessario para calcular % sobra e CMS real."),
        ("percentual_forragem",  "Percentual de forragem na dieta, de 0 a 100 (ex: 52.5)."),
        ("pct_ms_forragem",      "Percentual de materia seca da fracao forragem (ex: 35.0)."),
    ]:
        wi_row(r, label, desc, "F3F4F6"); r += 1

    wi.row_dimensions[r].height = 6; r += 1
    wi_header(r, "CALCULOS REALIZADOS AUTOMATICAMENTE", "1A6B3A"); r += 1
    for label, desc in [
        ("CMS/vaca",       "CMS/vaca = (Dieta fornecida - Sobra) x %MS_dieta / 100 / Num_vacas  (requer sobra)"),
        ("Eficiencia",     "Eficiencia = Leite/vaca / CMS/vaca"),
        ("% Sobra",        "% Sobra = Sobra / Dieta fornecida x 100"),
        ("Leite/vaca",     "Calculado automaticamente se voce fornecer producao total"),
        ("Prod. total",    "Calculado automaticamente se voce fornecer leite/vaca"),
    ]:
        wi_row(r, label, desc, "DCFCE7"); r += 1

    wi.row_dimensions[r].height = 6; r += 1
    wi_header(r, "DICAS E ERROS COMUNS", "D97706"); r += 1
    for tip in [
        ("CORRETO",  "Cada linha representa um lote em um dia especifico."),
        ("CORRETO",  "Use ponto (.) como separador decimal, nao virgula."),
        ("CORRETO",  "Apague as linhas de exemplo (6 a 17) antes de enviar."),
        ("CORRETO",  "O sistema aceita arquivos .xlsx e .csv."),
        ("ATENCAO",  "Nao altere os nomes das colunas na planilha."),
        ("ATENCAO",  "Campos em branco sao aceitos — o sistema apenas nao fara o calculo."),
        ("ATENCAO",  "Nao use formulas nas celulas — apenas valores numericos."),
    ]:
        bg = "FFFDE7" if tip[0] == "CORRETO" else "FDECEA"
        wi_row(r, tip[0], tip[1], bg); r += 1

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=template_fazenda.xlsx"})


# ──────────────────────────────────────────────
# Serve React SPA (deve ficar no final)
# ──────────────────────────────────────────────

_dist = os.path.join(os.path.dirname(__file__), '..', 'frontend', 'dist')
if os.path.isdir(_dist):
    from fastapi.staticfiles import StaticFiles
    from fastapi.responses import FileResponse

    _assets = os.path.join(_dist, 'assets')
    if os.path.isdir(_assets):
        app.mount('/assets', StaticFiles(directory=_assets), name='assets')

    @app.get('/{full_path:path}', include_in_schema=False)
    async def serve_spa(full_path: str):
        return FileResponse(os.path.join(_dist, 'index.html'))
